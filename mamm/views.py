# -*- coding: utf-8 -*-
from __future__ import unicode_literals

#Default
import requests
import openpyxl
import jwt

#Django
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.core import serializers
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.utils.crypto import get_random_string

#Rest_Framework
from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView
from rest_framework import permissions
from rest_framework import status

from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

#Own
from models import Patient, Stuff, MedicalHistory, MedicalHistoryExcel, MedicalHistoryExcelTemplate, Doctor, MedicineHistoryExcel, TreatmentHistoryExcel, ReferralHistory, PdfTemplate, TextImageTemplate
from serializers import PatientSerializer


# SECRET FOR JWT
SECRET = 'Y2UiOmZhbHNlLCJwaG9uZW'
# Create your views here.

class PatientViewSet(ModelViewSet):
    # this fetches all the rows of data in the Patient table
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer

def index(request):
    templates = MedicalHistoryExcelTemplate.objects.all()
    context = {
        'templates': templates
    }
    for template in templates:
        print(template.excel.path)
    print(templates)
    return render(request, 'mamm/medical/index.html', context)

"""
SMS Service
"""

def send_sms_to_stuff(patient):
    stuffs = Stuff.objects.all()
    send_string = "Name:" + patient.first_name +\
                 "\nPhone:" + patient.phonenumber

    print(send_string)
    for stuff in stuffs:

        payload = {
            'userid' : "",
            'account' : "jkwl518",
            'password' : "jkwl51825",
            'mobile' : stuff.phone,
            'content': "Patient Ordered Service!\nUser Detail\n" +send_string+ "【爱克】",
            'sendTime' : "",
            'extno' : ""
        }

        headers = {'content-type': 'application/x-www-form-urlencoded;charset=utf-8'}

        r = requests.post("https://sh2.ipyy.com/smsJson.aspx?action=send", data=payload, headers=headers)
        print(r.status_code)

"""
Get and Decode the Token in the Header
"""
def GetAndDecodeToken(request):
    if request.META['HTTP_AUTHORIZATION'] is not None:
        token = request.META['HTTP_AUTHORIZATION']
        token = token.replace('Bearer ', '')
        try:
            decoded_token = jwt.decode(token, SECRET, algorithms=['HS256'])
        except jwt.InvalidTokenError:
            return False
        except jwt.ExpiredSignature:
            return False
        except jwt.DecodeError:
            return False
        return decoded_token
    else:
        return False

def send_verify_sms(phonenumber, content):
    payload = {
        'userid' : "",
        'account' : "jkwl518",
        'password' : "jkwl51825",
        'mobile' : phonenumber,
        'content': content,
        'sendTime' : "",
        'extno' : ""
    }

    headers = {'content-type': 'application/x-www-form-urlencoded;charset=utf-8'}

    r = requests.post("https://sh2.ipyy.com/smsJson.aspx?action=send", data=payload, headers=headers)
    return r.status_code

def get_additional_data(patient):
    additional_data = {}
    if patient.disease_type != None and patient.disease_type.diseasetype != 'applogo':
        text_image = TextImageTemplate.objects.get(disease_type = patient.disease_type)
        if text_image.image1:
            additional_data['image1'] = text_image.image1.url
        if text_image.image2:
            additional_data['image2'] = text_image.image2.url
        if text_image.image3:
            additional_data['image3'] = text_image.image3.url
        if text_image.image4:
            additional_data['image4'] = text_image.image4.url
        if text_image.image5:
            additional_data['image5'] = text_image.image5.url
        if text_image.image6:
            additional_data['image6'] = text_image.image6.url
        if text_image.image7:
            additional_data['image7'] = text_image.image7.url
        if text_image.image8:
            additional_data['image8'] = text_image.image8.url
        if text_image.image9:
            additional_data['image9'] = text_image.image9.url
        if text_image.image10:
            additional_data['image10'] = text_image.image10.url
        if text_image.heading1:
            additional_data['heading1'] = text_image.heading1
        if text_image.heading2:
            additional_data['heading2'] = text_image.heading2
        if text_image.heading3:
            additional_data['heading3'] = text_image.heading3
        if text_image.heading4:
            additional_data['heading4'] = text_image.heading4
        if text_image.heading5:
            additional_data['heading5'] = text_image.heading5
        if text_image.heading6:
            additional_data['heading6'] = text_image.heading6
        if text_image.heading7:
            additional_data['heading7'] = text_image.heading7
        if text_image.heading8:
            additional_data['heading8'] = text_image.heading8
        if text_image.heading9:
            additional_data['heading9'] = text_image.heading9
        if text_image.heading10:
            additional_data['heading10'] = text_image.heading10
        if text_image.content1:
            additional_data['content1'] = text_image.content1
        if text_image.content2:
            additional_data['content2'] = text_image.content2
        if text_image.content3:
            additional_data['content3'] = text_image.content3
        if text_image.content4:
            additional_data['content4'] = text_image.content4
        if text_image.content5:
            additional_data['content5'] = text_image.content5
        if text_image.content6:
            additional_data['content6'] = text_image.content6
        if text_image.content7:
            additional_data['content7'] = text_image.content7
        if text_image.content8:
            additional_data['content8'] = text_image.content8
        if text_image.content9:
            additional_data['content9'] = text_image.content9
        if text_image.content10:
            additional_data['content10'] = text_image.content10
    return additional_data
       
@api_view(['POST'])
def login(request):
    if request.method == 'POST':
        email = request.data['email']
        password = request.data['password']
        
        try:
            patient = Patient.objects.get(email=email)
        except Patient.DoesNotExist:
            patient = None
        if patient is None:
            try:
                patient = Patient.objects.get(phonenumber=email)
            except Patient.DoesNotExist:
                patient = None

        if patient is not None:

            if check_password(password, patient.password):
                additional_data = get_additional_data(patient)
                send_data = {
                    'first_name': patient.first_name,
                    'last_name': patient.last_name,
                    'email': patient.email,
                    'phonenumber': patient.phonenumber,
                    'chujonservice': patient.chujonservice,
                    'huijonservice': patient.huijonservice,
                    'jiuyiservice': patient.jiuyiservice,
                    'additional_data': additional_data
                }

                print(additional_data)                    
                print(send_data)
                encoded_token = jwt.encode(send_data, SECRET, algorithm = 'HS256').decode('utf-8')
                print(encoded_token)
                return Response({ 'data': send_data, 'token': 'Bearer ' + encoded_token }, status=status.HTTP_201_CREATED)

            else:
                return Response({'error':{'message':"Password Incorrect!"}})
        else:
            return Response({'error':{'message':"User doesn't exist"}})


@api_view(['POST'])
def register(request):
    print("Register")
    if request.method == 'POST':
        email = request.data['email']
        phonenumber = request.data['phonenumber']
        referralvcode = request.data['referralvcode']
        if verifyvcode_referral(phonenumber, referralvcode) == 1:
            try:
                patient = Patient.objects.get(email=email)
            except Patient.DoesNotExist:
                patient = None
            
            if patient is not None:
                return Response({'error': { 'message':"User with that email already exist!"}})
            else:
                try:
                    patient = Patient.objects.get(phonenumber=phonenumber)
                except Patient.DoesNotExist:
                    patient = None

                if patient is None:
                    password = request.data['password']
                    print(make_password(password))
                    request.data['password'] = make_password(password)

                    serializer = PatientSerializer(data=request.data)
                    if serializer.is_valid():
                        serializer.save()
                        return Response({'data': {'message' : 'Success'}}, status=200)
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                    
                else:
                    return Response({'error': { 'message':"User with that phonenumber already exist!"}})
        else:
            return Response({'error': {'message': "VerifyCode Mismatch"}})


@api_view(['POST'])
def chuzonservice(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        email = decoded_token['email']

        try:
            patient = Patient.objects.get(email=email)
        except Patient.DoesNotExist:
            patient = None

        if patient is not None:
            # send_sms_to_stuff(patient)
            return Response({},status = 200)

        else:
            return Response({'error': {'message':"Can't find user"}}, status=400)
    else:
        return Response({'error': { 'message' : "Unidentified User"}}, status=400)



@api_view(['GET'])
def getmedicinehistory(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        print(decoded_token)
        historys = MedicineHistoryExcel.objects.all()
        send_data = []

        for history in historys:
            print(history.excel)
            book = openpyxl.load_workbook(history.excel)
            sheet = book.active
            dicti = {}

            dicti['patient_name'] = history.patient.first_name + ' ' + history.patient.last_name
            dicti['phonenumber'] = history.patient.phonenumber
            dicti['creation_date'] = history.creation_date

            excel_data = []
            for rowObjects in sheet[sheet.dimensions]:
                if rowObjects[1].value is not None:
                    excel_dictionary = {}
                    excel_dictionary[rowObjects[0].value] = rowObjects[1].value
                    excel_data.append(excel_dictionary)

            dicti['excel_data'] = excel_data
    
            if history.image1:
                print(history.image1.url)
                dicti['image1'] = history.image1.url

            if history.image2:
                dicti['image2'] = history.image2.url

            if history.video1:
                dicti['video1'] = history.video1.url
            send_data.append(dicti)

        return Response({'data': send_data },status = 200)

    else:
        return Response({'error': { 'message' : "Unidentified User"}}, status=400)

@api_view(['GET'])
def gettreatmenthistory(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        print(decoded_token)
        historys = TreatmentHistoryExcel.objects.all()
        send_data = []

        for history in historys:
            print(history.excel)
            book = openpyxl.load_workbook(history.excel)
            sheet = book.active
            dicti = {}

            dicti['patient_name'] = history.patient.first_name + ' ' + history.patient.last_name
            dicti['phonenumber'] = history.patient.phonenumber
            dicti['creation_date'] = history.creation_date

            excel_data = []
            for rowObjects in sheet[sheet.dimensions]:
                if rowObjects[1].value is not None:
                    excel_dictionary = {}
                    excel_dictionary[rowObjects[0].value] = rowObjects[1].value
                    excel_data.append(excel_dictionary)

            dicti['excel_data'] = excel_data
    
            if history.image1:
                print(history.image1.url)
                dicti['image1'] = history.image1.url

            if history.image2:
                dicti['image2'] = history.image2.url

            if history.video1:
                dicti['video1'] = history.video1.url
            send_data.append(dicti)

        return Response({'data': send_data },status = 200)

    else:
        return Response({'error': { 'message' : "Unidentified User"}}, status=400)

@api_view(['GET'])
def getmedicalhistory(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        print(decoded_token)
        historys = MedicalHistoryExcel.objects.all()
        send_data = []

        for history in historys:
            print(history.excel)
            book = openpyxl.load_workbook(history.excel)
            sheet = book.active
            dicti = {}

            dicti['patient_name'] = history.patient.first_name + ' ' + history.patient.last_name
            dicti['phonenumber'] = history.patient.phonenumber
            dicti['creation_date'] = history.creation_date
            
            '''
            Get the PDF
            '''
            dicti['pdf'] = ""
            if history.disease_type != None and history.disease_type.diseasetype != 'applogo':
                pdf_template = PdfTemplate.objects.get(disease_type=history.disease_type)
                dicti['pdf'] = pdf_template.pdf.url

            excel_data = []
            for rowObjects in sheet[sheet.dimensions]:
                if rowObjects[1].value is not None:
                    excel_dictionary = {}
                    excel_dictionary[rowObjects[0].value] = rowObjects[1].value
                    excel_data.append(excel_dictionary)

            dicti['excel_data'] = excel_data
    
            if history.image1:
                print(history.image1.url)
                dicti['image1'] = history.image1.url

            if history.image2:
                dicti['image2'] = history.image2.url

            if history.video1:
                dicti['video1'] = history.video1.url
            send_data.append(dicti)

        return Response({'data': send_data },status = 200)

    else:
        return Response({'error': { 'message' : "Unidentified User"}}, status=400)

@api_view(['POST'])
def forgotpass(request):
    print(request.data)

    try:
        phone = request.data['phone']
    except KeyError:
        return Response({'error':{'message':'KeyError'}})

    try:
        patient = Patient.objects.get(phonenumber=phone)
    except Patient.DoesNotExist:
        patient = None

    if patient is not None:
        rand_str = get_random_string(length=6, allowed_chars='1234567890')
        patient.verifycode = rand_str
        patient.save()
        content = "Verify Code is "+ rand_str +"【爱克】"
        statuscode = send_verify_sms(patient.phonenumber, content)
        return Response({'data': 'Success'}, status=200)

    return Response({'error' : {'message':'Patient does not exist'}})

@api_view(['POST'])
def verifycode(request):
    phone = request.data['phone']
    verifycode = request.data['verifycode']

    try:
        patient = Patient.objects.get(phonenumber=phone)
    except Patient.DoesNotExist:
        patient = None

    if patient is not None:
        if patient.verifycode == verifycode:
            return Response({'data':'Success'}, status=200)
        else:
            return Response({'error':{'message':'Wrong Verify Code'}})

    return Response({'error': {'message' : 'Failed'}})

@api_view(['POST'])
def setpass(request):
    phone = request.data['phone']
    password = request.data['password']

    try:
        patient = Patient.objects.get(phonenumber=phone)
    except Patient.DoesNotExist:
        patient = None

    if patient is not None:
        patient.password = make_password(password)
        patient.save()
        return Response({'data':'Success'}, status= 200)

    return Response({'error': { 'message': 'Failed'}})


"""
Login with Phone

"""
@api_view(['POST'])
def loginphone(request):
    try:
        phone = request.data['phone']
    except KeyError:
        return Response({'error':{'message':'KeyError'}})
    
    try:
        patient = Patient.objects.get(phonenumber=phone)
    except Patient.DoesNotExist:
        patient = None
    
    if patient is None:
        patient = Patient.objects.create(phonenumber=phone)
    
    rand_str = get_random_string(length=6, allowed_chars='1234567890')
    print(rand_str)
    patient.verifycode = rand_str
    patient.save()
    content = "【用心医】欢迎您，您的验证码是 "+ rand_str +"【爱克】"
    status_code = send_verify_sms(patient.phonenumber, content)

    if status_code == 200:
        send_data = {
            'phonenumber': patient.phonenumber,
            'chujonservice': patient.chujonservice,
            'huijonservice': patient.huijonservice,
            'jiuyiservice': patient.jiuyiservice
        }
        return Response({ 'data' : 'Success' }, status=200)
    else:
        return Response({'error': {'message': 'Failed'}}, status=status_code)

@api_view(['POST'])
def verifycode_phone(request):
    phone = request.data['phone']
    verifycode = request.data['verifycode']

    try:
        patient = Patient.objects.get(phonenumber=phone)
    except Patient.DoesNotExist:
        patient = None

    if patient is not None:
        if patient.verifycode == verifycode:
            send_data = {
                'phonenumber': patient.phonenumber,
                'chujonservice': patient.chujonservice,
                'huijonservice': patient.huijonservice,
                'jiuyiservice': patient.jiuyiservice
            }
            print(send_data)
            encoded_token = jwt.encode(send_data, SECRET, algorithm = 'HS256').decode('utf-8')
            print(encoded_token)
            return Response({ 'data': send_data, 'token':'Bearer ' + encoded_token }, status=200)
        else:
            return Response({'error':{'message':'Wrong Verify Code'}})

    return Response({'error': {'message' : 'Failed'}})

@api_view(['GET'])
def getdoctors(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        req_type = ['huizon', 'jiuyi']
        type_id = request.GET.get('type', '')

        doctors = Doctor.objects.all()
        send_data = []
        for doctor in doctors:
            if doctor.doctor_type == req_type[int(type_id)]:
                print(doctor)
                dicti = {}

                if doctor.avatar:
                    dicti['avatar'] = doctor.avatar.url
                dicti['description'] = doctor.description
                dicti['first_name'] = doctor.first_name
                dicti['last_name'] = doctor.last_name
                dicti['doctor_type'] = doctor.doctor_type
                send_data.append(dicti)
        print(send_data)
        return Response({'data': send_data },status = 200)
    else:
        return Response({'error': { 'message' : "Unidentified User"}}, status=400)

@api_view(['POST'])
def sendvcode_referral(request):
    decoded_token = GetAndDecodeToken(request)
    if decoded_token != False:
        phonenumber = request.data['referralphone']
        rand_str = get_random_string(length=6, allowed_chars='1234567890')
        ref_history = ReferralHistory(phonenumber=phonenumber, verifycode=rand_str)
        ref_history.save()
        content = "【用心医】欢迎您，App程序下载地址是 " + "https://medicalapp" + "邀请码是"+ rand_str +"【爱克】"
        statuscode = send_verify_sms(phonenumber, content)
        return Response(status=200)

def verifyvcode_referral(phonenumber, vcode):
    try:
        referral_history = ReferralHistory.objects.get(phonenumber = phonenumber)
    except ReferralHistory.DoesNotExist:
        return 2 
    if referral_history.verifycode == vcode:
        return 1
    else:
        return 0

@api_view(['GET'])
def getpdftemplate(request, slug):
    print(slug)
    try:
        pdftemplate = PdfTemplate.objects.get(description=slug)
    except PdfTemplate.DoesNotExist:
        return Response({'error': { 'message' : "No Pdf template"}}, status=400)
    return Response({'data': pdftemplate.pdf.url},status = 200)
    

@staff_member_required
def sendbulkmessage(request):
    return render(request, 'mamm/medical/bulkmessage.html')

@login_required
@api_view(['POST'])
def processbulkmessage(request):
    invalid_phonenumber = 0
    registered_phonenumber = 0
    count = request.POST.get('savecount')
    oldpeople = request.POST.get('oldpeople')
    print(count)
    print(oldpeople)
    if oldpeople == 'True':
        if request.user.is_superuser:
            for i in range(0, int(count)):
                # print(request.POST.get('phonenumber' +`i`))
                phonenumber = request.POST.get('phonenumber' +`i`)
                if phonenumber != None and len(phonenumber) >= 8:
                    rand_str = phonenumber[-6:]
                    ref_history = ReferralHistory(phonenumber=phonenumber, verifycode=rand_str)
                    ref_history.save()
                    try:
                        patient = Patient.objects.get(phonenumber=phonenumber)
                    except Patient.DoesNotExist:
                        patient = None
                    if patient is None:
                        patient = Patient()
                        patient.phonenumber = phonenumber
                        patient.password = make_password(rand_str)
                        patient.save()
                        content = "【用心医】欢迎您，App程序下载地址是 " + "https://medicalapp" + "用户名是您的手机号码，您的默认登陆密码为手机号码的后6位数字"+ "【爱克】"
                        statuscode = send_verify_sms(phonenumber, content)
                    else:
                        registered_phonenumber = registered_phonenumber + 1
                else:
                    invalid_phonenumber = invalid_phonenumber + 1
            
            if invalid_phonenumber != 0:
                return HttpResponse("Message Sent, but there are " + str(invalid_phonenumber) + " invalid phonenumbers and " + str(registered_phonenumber) + "Registered Phonenumbers.")
            return HttpResponse("Message Sent")
        else:
            return HttpResponse("工作人员不能操作")

    for i in range(0, int(count)):
        # print(request.POST.get('phonenumber' +`i`))
        phonenumber = request.POST.get('phonenumber' +`i`)
        if phonenumber != None and len(phonenumber) > 8 :
            rand_str = get_random_string(length=6, allowed_chars='1234567890')
            ref_history = ReferralHistory(phonenumber=phonenumber, verifycode=rand_str)
            ref_history.save()
            content = "【用心医】欢迎您，App程序下载地址是 " + "https://medicalapp" + "邀请码是"+ rand_str +"【爱克】"
            statuscode = send_verify_sms(phonenumber, content)
        else:
            invalid_phonenumber = invalid_phonenumber + 1
    
    if invalid_phonenumber != 0:
        return HttpResponse("Message Sent, but there are " + str(invalid_phonenumber) + " invalid phonenumbers")
    return HttpResponse("Message Sent")